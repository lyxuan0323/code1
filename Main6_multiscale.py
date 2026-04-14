# -*- coding: utf-8 -*-

"""
Main6_multiscale

This module estimates building heights from panoramic imagery and building masks,
and evaluates the computational efficiency of the fixed-step search and the
coarse-to-fine multiscale search.

In this version, the single-corner height estimation procedure has been rewritten
to match the textual logic presented in Section 3.3.1 of the paper:

1) The bottom point of each building corner is first projected onto the panoramic
   image to determine the fixed column coordinate u_i associated with that corner.
2) The building bottom boundary b(u_i) and roof boundary r(u_i) are then read
   from that fixed column.
3) Candidate heights are searched along the vertical direction of the corner.
4) For each candidate height h, only the relationship between its projected row
   coordinate v_i(h) and the roof boundary r(u_i) in the same column is evaluated.
5) When v_i(h) first reaches or crosses the roof boundary, that is, when
   v_i(h) <= r(u_i), the corresponding h is recorded as the estimated height
   of that corner.
6) On this basis, two search strategies are implemented:
   - fixed_step: point-by-point accumulation with a fixed step of 0.1 m
   - multiscale: progressive coarse-to-fine refinement with step sizes of
     10, 5, 2.5, 1, 0.5, and 0.1 m

Compared with the original implementation, the key changes in this version are
as follows:

1) The projected column of the bottom point is used as the fixed search column,
   so the roof boundary is not repeatedly queried from the current projected
   column of each candidate point during the search process.
2) The bottom boundary is explicitly used as the starting reference in image space.
3) The search starts from 0 m relative to the building base, which is more
   consistent with the definition of height as measured from the bottom upward.
4) The matching condition for each corner is uniformly defined as
   v_i(h) <= r(u_i).
"""
import cv2
import numpy as np
import project_pano as p2pano
import pandas as pd
from openpyxl import load_workbook
import os
import json
import time
from statistics import mean


# ==============================
# 0. Global parameters
# ==============================
CAMERA_HEIGHT_M = 2.5
HEIGHT_MIN_M = 0.0
HEIGHT_MAX_M = 100.0
FIXED_STEP_M = 0.1
MULTISCALE_STEPS_M = [10.0, 5.0, 2.5, 1.0, 0.5, 0.1]
COLUMN_TOLERANCE_PX = 2

# Whether to run the search strategy evaluation
RUN_STRATEGY_EXPERIMENT = True
# Number of rectified panoramas used in the evaluation; set to 1 to process one sample dataset
EXPERIMENT_MAX_IMAGES = 1
# Output file
STRATEGY_EVAL_CSV = "search_strategy_evaluation.csv"


# ==============================
# 1. Excel result appending function
# ==============================
def append_nested_dict_to_excel(file_path, sheet_name, nested_dict):
    rows = []
    for key, value in nested_dict.items():
        row = {'Index': key}
        row.update(value)
        rows.append(row)

    df_new = pd.DataFrame(rows)
    df_new = df_new.set_index('Index')

    if not os.path.exists(file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_new.to_excel(writer, sheet_name=sheet_name, index=True)
    else:
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            df_existing = pd.read_excel(file_path, sheet_name=sheet_name, index_col=0)
            df_combined = pd.concat([df_existing, df_new], axis=1)
        else:
            df_combined = df_new

        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_combined.to_excel(writer, sheet_name=sheet_name, index=True)


# ==============================
# 2. Extract top and bottom boundaries from the binary building mask
# ==============================
def getbldbound(mask_path):
    mask = cv2.imread(mask_path, cv2.IMREAD_GRAYSCALE)
    if mask is None:
        raise FileNotFoundError(f"Unable to read the building mask：{mask_path}")

    kernel = np.ones((3, 3), np.uint8)
    mask_clean = cv2.morphologyEx(mask, cv2.MORPH_OPEN, kernel)

    h, w = mask_clean.shape

    top_boundary_points = np.zeros((w, 2), dtype=int)
    bottom_boundary_points = np.zeros((w, 2), dtype=int)

    for u in range(w):
        column = mask_clean[:, u]
        ys = np.where(column > 0)[0]

        if len(ys) > 0:
            v_top = int(ys[0])
            v_bottom = int(ys[-1])
        else:
            v_top = 0
            v_bottom = 0

        top_boundary_points[u] = [u, v_top]
        bottom_boundary_points[u] = [u, v_bottom]

    vis_img = cv2.cvtColor(mask_clean, cv2.COLOR_GRAY2BGR)
    for u in range(w):
        vt = top_boundary_points[u][1]
        vb = bottom_boundary_points[u][1]
        if vt > 0:
            vis_img[vt, u] = (0, 0, 255)   # roof boundary - red
        if vb > 0:
            vis_img[vb, u] = (255, 0, 0)   # bottom boundary - blue

    return top_boundary_points, bottom_boundary_points, vis_img


# ==============================
# 3. Projection helper functions
# ==============================
def project_point_to_pano(point_geo, camera_xyz, pano_width, pano_height,
                          north_rotation, camera_bearing):
    x, y, z = p2pano.geo_to_utm(*point_geo)
    camera_x, camera_y, camera_z = camera_xyz
    x_relative = x - camera_x
    y_relative = y - camera_y
    z_relative = z - camera_z
    u, v = p2pano.project_to_pano(
        x_relative, y_relative, z_relative,
        pano_width, pano_height,
        north_rotation, camera_bearing
    )
    return u, v


def get_corner_column_reference(point_xy, footheight, camera_xyz,
                                pano_width, pano_height,
                                north_rotation, camera_bearing,
                                top_boundary_points, bottom_boundary_points):

    point_base = [point_xy[0], point_xy[1], footheight]
    base_u, base_v = project_point_to_pano(
        point_base, camera_xyz,
        pano_width, pano_height,
        north_rotation, camera_bearing
    )
    ref_u_idx = int(round(base_u))

    if not (0 <= ref_u_idx < pano_width):
        return None, base_u, base_v, None, None, False

    bottom_v = int(bottom_boundary_points[ref_u_idx][1])
    top_v = int(top_boundary_points[ref_u_idx][1])

    # If no valid building boundaries are available in this column, skip height estimation for this corner
    if bottom_v <= 0 or top_v <= 0 or bottom_v <= top_v:
        return ref_u_idx, base_u, base_v, bottom_v, top_v, False

    return ref_u_idx, base_u, base_v, bottom_v, top_v, True


def evaluate_candidate_height_same_column(point_xy, candidate_h, camera_xyz,
                                          pano_width, pano_height,
                                          north_rotation, camera_bearing,
                                          ref_u_idx, roof_v):

    point_geo = [point_xy[0], point_xy[1], candidate_h]
    u2, v2 = project_point_to_pano(
        point_geo, camera_xyz,
        pano_width, pano_height,
        north_rotation, camera_bearing
    )

    if not np.isfinite(u2) or not np.isfinite(v2):
        return False, False, u2, v2, None, False

    same_column = abs(int(round(u2)) - ref_u_idx) <= COLUMN_TOLERANCE_PX
    residual = v2 - roof_v
    hit = residual <= 0
    return hit, True, u2, v2, residual, same_column


# ==============================
# 4. Two height search strategies
# ==============================
def search_height_fixed_step(point_xy, footheight, camera_xyz,
                             pano_width, pano_height,
                             north_rotation, camera_bearing,
                             ref_u_idx, roof_v,
                             h_min=HEIGHT_MIN_M, h_max=HEIGHT_MAX_M,
                             step=FIXED_STEP_M):

    projection_count = 0
    est_h = None
    hit_u, hit_v = None, None

    h = h_min
    while h <= h_max + 1e-9:
        projection_count += 1
        hit, valid, u2, v2, residual, same_column = evaluate_candidate_height_same_column(
            point_xy=point_xy,
            candidate_h=footheight + h,
            camera_xyz=camera_xyz,
            pano_width=pano_width,
            pano_height=pano_height,
            north_rotation=north_rotation,
            camera_bearing=camera_bearing,
            ref_u_idx=ref_u_idx,
            roof_v=roof_v,
        )

        if not valid:
            h += step
            continue

        if hit:
            est_h = round(h, 1)
            hit_u, hit_v = u2, v2
            break

        h += step

    return est_h, projection_count, hit_u, hit_v


def search_height_multiscale(point_xy, footheight, camera_xyz,
                             pano_width, pano_height,
                             north_rotation, camera_bearing,
                             ref_u_idx, roof_v,
                             h_min=HEIGHT_MIN_M, h_max=HEIGHT_MAX_M,
                             step_schedule=MULTISCALE_STEPS_M):

    projection_count = 0
    est_h = None
    hit_u, hit_v = None, None

    low = h_min
    high = h_max

    for step in step_schedule:
        found = False
        h = low

        while h <= high + 1e-9:
            projection_count += 1
            hit, valid, u2, v2, residual, same_column = evaluate_candidate_height_same_column(
                point_xy=point_xy,
                candidate_h=footheight + h,
                camera_xyz=camera_xyz,
                pano_width=pano_width,
                pano_height=pano_height,
                north_rotation=north_rotation,
                camera_bearing=camera_bearing,
                ref_u_idx=ref_u_idx,
                roof_v=roof_v,
            )

            if not valid:
                h += step
                continue

            if hit:
                est_h = h
                hit_u, hit_v = u2, v2
                low = max(h_min, h - step)
                high = h
                found = True
                break

            h += step

        if not found:
            return None, projection_count, None, None

    if est_h is not None:
        est_h = round(est_h, 1)

    return est_h, projection_count, hit_u, hit_v


# ==============================
# 5. Height iteration function
# ==============================
def iterheight(pano_image_path, points_geo, camera_geo, pano_size,
               north_rotation, camera_bearing, footheight,
               top_boundary_points, bottom_boundary_points,
               search_mode="multiscale", optimal=None):

    pano_img = cv2.imread(pano_image_path)
    if pano_img is None:
        print("Unable to load the panoramic image：", pano_image_path)
        return {}, {}

    pano_height, pano_width = pano_size
    camera_xyz = p2pano.geo_to_utm(*camera_geo)

    bldheightset = {}
    per_corner_projection_counts = []
    valid_corner_num = 0

    t0 = time.perf_counter()

    for index in points_geo.keys():
        if index in bldheightset:
            continue

        height = []

        for point in points_geo[index]:
            # ======= Establish the fixed column reference: u_i, b(u_i), r(u_i) =======
            ref_u_idx, base_u, base_v, bottom_v, roof_v, valid_ref = get_corner_column_reference(
                point_xy=point,
                footheight=footheight,
                camera_xyz=camera_xyz,
                pano_width=pano_width,
                pano_height=pano_height,
                north_rotation=north_rotation,
                camera_bearing=camera_bearing,
                top_boundary_points=top_boundary_points,
                bottom_boundary_points=bottom_boundary_points,
            )

            if not valid_ref:
                continue

            bottom_x = ref_u_idx
            bottom_y = int(bottom_v)

            # Visualization: reference point on the bottom boundary
            if 0 <= bottom_x < pano_width and 0 <= bottom_y < pano_height:
                cv2.circle(pano_img, (bottom_x, bottom_y), 5, (0, 0, 255), -1)

            # ======= Top search: always compare with roof_v in the same column =======
            if search_mode == "fixed":
                est_h, proj_count, hit_u, hit_v = search_height_fixed_step(
                    point_xy=point,
                    footheight=footheight,
                    camera_xyz=camera_xyz,
                    pano_width=pano_width,
                    pano_height=pano_height,
                    north_rotation=north_rotation,
                    camera_bearing=camera_bearing,
                    ref_u_idx=ref_u_idx,
                    roof_v=roof_v,
                )
            else:
                est_h, proj_count, hit_u, hit_v = search_height_multiscale(
                    point_xy=point,
                    footheight=footheight,
                    camera_xyz=camera_xyz,
                    pano_width=pano_width,
                    pano_height=pano_height,
                    north_rotation=north_rotation,
                    camera_bearing=camera_bearing,
                    ref_u_idx=ref_u_idx,
                    roof_v=roof_v,
                )

            per_corner_projection_counts.append(proj_count)

            if est_h is not None:
                height.append(est_h)
                valid_corner_num += 1

                if hit_u is not None and hit_v is not None:
                    hit_u_i = int(round(hit_u))
                    hit_v_i = int(round(hit_v))
                    if 0 <= hit_u_i < pano_width and 0 <= hit_v_i < pano_height:
                        cv2.circle(pano_img, (hit_u_i, hit_v_i), 5, (255, 0, 0), -1)
                        cv2.line(pano_img, (bottom_x, bottom_y), (hit_u_i, hit_v_i), (255, 0, 0), 3)

        bldheight = sum(height) / len(height) if len(height) > 0 else 0
        bldheightset[index] = {'bldheight': bldheight, 'height': height}

    elapsed = time.perf_counter() - t0

    os.makedirs("image", exist_ok=True)
    base_name = os.path.splitext(os.path.basename(pano_image_path))[0]
    out_vis = os.path.join("image", f"marked_iterheight_{search_mode}_{base_name}.jpg")
    cv2.imwrite(out_vis, pano_img)
    print(f"✔ Saved {search_mode} visualization image：", out_vis)

    stats = {
        "search_mode": search_mode,
        "image_name": os.path.basename(pano_image_path),
        "valid_corner_num": valid_corner_num,
        "avg_projections_per_corner": float(mean(per_corner_projection_counts)) if per_corner_projection_counts else 0.0,
        "max_projections_per_corner": int(max(per_corner_projection_counts)) if per_corner_projection_counts else 0,
        "min_projections_per_corner": int(min(per_corner_projection_counts)) if per_corner_projection_counts else 0,
        "elapsed_seconds": elapsed
    }

    return bldheightset, stats


# ==============================
# 6. Batch matching of raw / rect / mask files
# ==============================
def read_files_from_folders(base_folder):
    rawdata_folder = os.path.join(base_folder, 'rawdata')
    rectdata_folder = os.path.join(base_folder, 'rectdata')
    svgdata_folder = os.path.join(base_folder, 'svgdata')

    rawdata_files = set(os.listdir(rawdata_folder))
    rectdata_files = set(os.listdir(rectdata_folder))
    svgdata_files = set(os.listdir(svgdata_folder))

    RawDataFile, RectDataFile, SVGDataFile = [], [], []

    for raw_file in rawdata_files:
        base_name = os.path.splitext(raw_file)[0]
        ext = os.path.splitext(raw_file)[1]

        rect_file = f"rectified_panorama_{base_name}{ext}"
        svg_file = f"building_mask_{base_name}{ext}"

        if rect_file in rectdata_files and svg_file in svgdata_files:
            print(f"Raw Data File: {os.path.join(rawdata_folder, raw_file)}")
            print(f"Rect Data File: {os.path.join(rectdata_folder, rect_file)}")
            print(f"SVG/Mask Data File: {os.path.join(svgdata_folder, svg_file)}")
            print("-" * 40)
            RawDataFile.append(os.path.join(rawdata_folder, raw_file))
            RectDataFile.append(os.path.join(rectdata_folder, rect_file))
            SVGDataFile.append(os.path.join(svgdata_folder, svg_file))

    return RawDataFile, RectDataFile, SVGDataFile


# ==============================
# 7. Load metadata JSON files
# ==============================
def load_metadata_jsons(json_folder):
    meta_dict = {}

    if not os.path.exists(json_folder):
        print("⚠ Metadata directory does not exist：", json_folder)
        return meta_dict

    for file in os.listdir(json_folder):
        if file.endswith(".metadata.json"):
            num = ''.join(filter(str.isdigit, file))
            if not num:
                continue

            json_path = os.path.join(json_folder, file)
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            meta_dict[num] = {
                "lat": data["lat"],
                "lng": data["lng"],
                "elevation": data.get("elevation", 1.7),
                "rotation": data.get("rotation", 0.0)
            }

    print("✔ Number of metadata files loaded：", len(meta_dict))
    return meta_dict


# ==============================
# 8. Save strategy evaluation results
# ==============================
def save_strategy_eval_csv(basefolder, rows):
    if not rows:
        return
    out_csv = os.path.join(basefolder, STRATEGY_EVAL_CSV)
    pd.DataFrame(rows).to_csv(out_csv, index=False, encoding="utf-8-sig")
    print("✔ Saved Search Strategy Evaluation：", out_csv)


# ==============================
# 9. Main program
# ==============================
if __name__ == "__main__":

    # ===== 1. Basic path settings =====
    basefolder = r"F:\gsv3d\code1\gsi\area"
    json_folder = os.path.join(basefolder, "360json")
    osmfile = r'F:\gsv3d\code1\osm\rectdata.shp'

    # ===== 2. Load metadata =====
    meta_dict = load_metadata_jsons(json_folder)

    # ===== 3. Load the raw / rect / mask file list =====
    RawDataFile, RectDataFile, SVGDataFile = read_files_from_folders(basefolder)

    strategy_eval_rows = []
    processed_n = 0

    # ===== 4. Batch process each rectified panorama =====
    for ind in range(len(RectDataFile)):

        rectd = RectDataFile[ind]
        maskd = SVGDataFile[ind]

        rect_filename = os.path.basename(rectd)
        rect_name_no_ext = os.path.splitext(rect_filename)[0]
        rect_num = ''.join(filter(str.isdigit, rect_filename))

        if not rect_num:
            print("⚠ Unable to extract an ID from the filename; skipping：", rect_filename)
            continue

        if rect_num not in meta_dict:
            print("❌ ID not found in metadata：", rect_num, " file：", rect_filename)
            continue

        cam = meta_dict[rect_num]

        camera_geo = [cam["lng"], cam["lat"], CAMERA_HEIGHT_M]
        north_rotation = cam["rotation"]
        camera_bearing = 180 - north_rotation

        pano_img = cv2.imread(rectd)
        if pano_img is None:
            print("❌ Unable to read the rectified image：", rectd)
            continue

        pano_size = pano_img.shape[:2]
        points_geo = p2pano.select_footprint(osmfile, camera_geo)[1]
        top_boundary_points, bottom_boundary_points, boundary_image = getbldbound(maskd)

        # ===== Default output: multiscale search results =====
        bldheightset, stats_multiscale = iterheight(
            rectd, points_geo, camera_geo, pano_size,
            north_rotation, camera_bearing,
            footheight=0,
            top_boundary_points=top_boundary_points,
            bottom_boundary_points=bottom_boundary_points,
            search_mode="multiscale"
        )

        if not bldheightset:
            print("⚠ No height result was obtained; skipping save：", rect_filename)
            continue

        out_excel_name = f"bldheightset_multiscale_{rect_name_no_ext}.xlsx"
        out_excel_path = os.path.join(basefolder, out_excel_name)
        append_nested_dict_to_excel(out_excel_path, "Sheet1", bldheightset)
        print("✔ Saved height results：", out_excel_path)

        # ===== Strategy evaluation: fixed-step vs multiscale =====
        if RUN_STRATEGY_EXPERIMENT and processed_n < EXPERIMENT_MAX_IMAGES:
            _, stats_fixed = iterheight(
                rectd, points_geo, camera_geo, pano_size,
                north_rotation, camera_bearing,
                footheight=0,
                top_boundary_points=top_boundary_points,
                bottom_boundary_points=bottom_boundary_points,
                search_mode="fixed"
            )

            strategy_eval_rows.append({
                "image_name": rect_filename,
                "fixed_avg_proj_per_corner": stats_fixed["avg_projections_per_corner"],
                "fixed_elapsed_s": stats_fixed["elapsed_seconds"],
                "fixed_valid_corner_n": stats_fixed["valid_corner_num"],
                "multiscale_avg_proj_per_corner": stats_multiscale["avg_projections_per_corner"],
                "multiscale_elapsed_s": stats_multiscale["elapsed_seconds"],
                "multiscale_valid_corner_n": stats_multiscale["valid_corner_num"],
                "final_height_resolution_m": 0.1
            })

            processed_n += 1

    if RUN_STRATEGY_EXPERIMENT:
        save_strategy_eval_csv(basefolder, strategy_eval_rows)
